#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Financial Model Generator for Day 24 Lab
Creates Excel file with 3 sheets: Assumptions, Unit Economics, P&L & ROI
Author: Bùi Thị Như Ngọc (2A202601882)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
from pathlib import Path

# Constants
FILE_NAME = "BuiThiNhuNgoc_Day24.xlsx"
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

BLUE_FONT = Font(name="Arial", size=10, color="0000FF", bold=False)
BLACK_FONT = Font(name="Arial", size=10, color="000000", bold=False)
GREEN_FONT = Font(name="Arial", size=10, color="008000", bold=False)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# Currency number format
CURRENCY_FORMAT = '#,##0;(#,##0);-'
PERCENT_FORMAT = '0.0%'
DECIMAL_FORMAT = '0.00'

def create_assumptions_sheet(ws):
    """Create Tab 1: Assumptions"""
    ws.title = "1. Assumptions"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 45
    
    # Header
    headers = ["", "Optimistic", "Base", "Pessimistic", "Unit", "Căn cứ / Benchmark"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Data structure: (Section Title, Row_offset, Data)
    # Each data item: (Label, Opt_val, Base_val, Pess_val, Unit, Benchmark)
    row = 2
    
    # Section 1: Product & Pricing
    ws.cell(row=row, column=1, value="1. PRODUCT & PRICING")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_product = [
        ("ARPU (VNĐ/khách/tháng)", 100, 80, 50, "M", "Salesforce Einstein 50-150M/tháng Enterprise VN"),
        ("Adoption Rate", 0.20, 0.15, 0.08, "%/tháng", "B2B SaaS Enterprise 0.08-0.2%/tháng benchmark"),
        ("TAM (tổng khách)", 650, 650, 650, "công ty", "650 Enterprise companies VN (Phase 0 calculation)"),
        ("Customers Month 0", 2, 1, 1, "khách", "Seed stage: 1-2 early pilot customers"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_product:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = DECIMAL_FORMAT
        row += 1
    
    # Section 2: COGS
    ws.cell(row=row, column=1, value="2. COGS / KHÁCH / THÁNG")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_cogs = [
        ("Model API Cost", 3.0, 4.0, 5.5, "M", "GPT-4 + Embeddings: 500 calls/tháng × 8k/call"),
        ("Data Labeling & QA", 1.0, 1.3, 2.0, "M", "Sales data labeling + feedback loop"),
        ("Model Retraining (20%/năm)", 0.8, 1.2, 2.0, "M", "Fine-tuning monthly per customer cohort"),
        ("Human-in-the-loop QA", 0.6, 1.0, 1.5, "M", "AI coaching output review by trainers"),
        ("Compliance & Security", 0.4, 0.5, 1.0, "M", "PDPA audit, SOC 2, data residency"),
        ("Infrastructure", 1.5, 2.0, 3.0, "M", "Multi-region deployment, 99.9% SLA"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_cogs:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = DECIMAL_FORMAT
        row += 1
    
    # Total COGS (formula)
    total_cogs_row = row
    ws.cell(row=row, column=1, value="Total COGS / khách / tháng")
    ws.cell(row=row, column=1).font = BOLD_FONT
    # Sum rows (API + Hidden 4 items + Infra)
    opt_total_cogs_row = row - 7  # First API Cost row
    ws.cell(row=row, column=2, value=f"=SUM(B{opt_total_cogs_row}:B{row-1})")
    ws.cell(row=row, column=2).font = GREEN_FONT
    ws.cell(row=row, column=3, value=f"=SUM(C{opt_total_cogs_row}:C{row-1})")
    ws.cell(row=row, column=3).font = GREEN_FONT
    ws.cell(row=row, column=4, value=f"=SUM(D{opt_total_cogs_row}:D{row-1})")
    ws.cell(row=row, column=4).font = GREEN_FONT
    ws.cell(row=row, column=5, value="M")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = DECIMAL_FORMAT
    row += 2
    
    # Section 3: Customer Behavior
    ws.cell(row=row, column=1, value="3. CUSTOMER BEHAVIOR")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_behavior = [
        ("Monthly Churn Rate", 0.01, 0.015, 0.0225, "%/tháng", "Enterprise SaaS 1-2%/tháng. Pess = Base × 1.5"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_behavior:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = PERCENT_FORMAT
        row += 1
    
    row += 1
    
    # Section 4: Sales & Marketing
    ws.cell(row=row, column=1, value="4. SALES & MARKETING")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_sales = [
        ("CAC (Customer Acquisition Cost)", 120, 150, 225, "M", "Enterprise sales: 2 AE × 6-9 month cycle"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_sales:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
        row += 1
    
    row += 1
    
    # Section 5: Fixed Costs
    ws.cell(row=row, column=1, value="5. FIXED COSTS / THÁNG")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_fixed = [
        ("Salaries (8 người)", 250, 300, 300, "M", "Founder + 2 AE + Sales Ops + 2 Eng + PM + CS"),
        ("Office & Tools", 50, 70, 70, "M", "Premium office space + SaaS tools"),
        ("Marketing Budget", 40, 60, 60, "M", "Content, events, ABM marketing"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_fixed:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
        row += 1
    
    # Total Fixed Costs
    total_fixed_row = row
    fixed_start_row = row - 3
    ws.cell(row=row, column=1, value="Total Fixed Costs / tháng")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"=SUM(B{fixed_start_row}:B{row-1})")
    ws.cell(row=row, column=2).font = GREEN_FONT
    ws.cell(row=row, column=3, value=f"=SUM(C{fixed_start_row}:C{row-1})")
    ws.cell(row=row, column=3).font = GREEN_FONT
    ws.cell(row=row, column=4, value=f"=SUM(D{fixed_start_row}:D{row-1})")
    ws.cell(row=row, column=4).font = GREEN_FONT
    ws.cell(row=row, column=5, value="M")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
    
    row += 2
    
    # Section 6: Investment & Cash
    ws.cell(row=row, column=1, value="6. INVESTMENT & CASH")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    data_invest = [
        ("Initial Investment (Month 0)", 1200, 1500, 1500, "M", "Engineering 800M + Infra 300M + Marketing 400M"),
        ("Initial Cash (runway buffer)", 4000, 5000, 3000, "M", "Series Seed funding round"),
        ("Discount Rate", 0.15, 0.25, 0.40, "%/năm", "SaaS startup valuation discount rate"),
    ]
    
    for label, opt, base, pess, unit, benchmark in data_invest:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=opt)
        ws.cell(row=row, column=2).fill = YELLOW_FILL
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=3).fill = YELLOW_FILL
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=4, value=pess)
        ws.cell(row=row, column=4).fill = YELLOW_FILL
        ws.cell(row=row, column=4).font = BLUE_FONT
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=benchmark)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col in [2, 3, 4]:
                if "Rate" in label:
                    ws.cell(row=row, column=col).number_format = PERCENT_FORMAT
                else:
                    ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
        row += 1
    
    row += 2
    
    # Checks Section
    ws.cell(row=row, column=1, value="RÀNG BUỘC KIỂM TRA")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    # Hidden Costs / API (should be ≥ 30%)
    hidden_start = 6  # Row of "Data Labeling"
    hidden_end = 9    # Row of "Compliance"
    api_row = 5       # Row of "Model API Cost"
    
    ws.cell(row=row, column=1, value="Hidden Costs / API Cost %")
    ws.cell(row=row, column=2, value=f"=(SUM(B{hidden_start}:B{hidden_end}))/B{api_row}")
    ws.cell(row=row, column=2).font = GREEN_FONT
    ws.cell(row=row, column=2).number_format = PERCENT_FORMAT
    ws.cell(row=row, column=3, value=f"=(SUM(C{hidden_start}:C{hidden_end}))/C{api_row}")
    ws.cell(row=row, column=3).font = GREEN_FONT
    ws.cell(row=row, column=3).number_format = PERCENT_FORMAT
    ws.cell(row=row, column=4, value=f"=(SUM(D{hidden_start}:D{hidden_end}))/D{api_row}")
    ws.cell(row=row, column=4).font = GREEN_FONT
    ws.cell(row=row, column=4).number_format = PERCENT_FORMAT
    ws.cell(row=row, column=5, value="(should ≥ 30%)")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    # Pess Churn / Base Churn (should be 1.5x)
    churn_row = 22  # Row of "Monthly Churn Rate"
    ws.cell(row=row, column=1, value="Pess Churn / Base Churn")
    ws.cell(row=row, column=3, value="(reference)")
    ws.cell(row=row, column=4, value=f"=D{churn_row}/C{churn_row}")
    ws.cell(row=row, column=4).font = GREEN_FONT
    ws.cell(row=row, column=4).number_format = DECIMAL_FORMAT
    ws.cell(row=row, column=5, value="(should = 1.5x)")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    # Pess CAC / Base CAC (should be 1.5x)
    cac_row = 25  # Row of "CAC"
    ws.cell(row=row, column=1, value="Pess CAC / Base CAC")
    ws.cell(row=row, column=3, value="(reference)")
    ws.cell(row=row, column=4, value=f"=D{cac_row}/C{cac_row}")
    ws.cell(row=row, column=4).font = GREEN_FONT
    ws.cell(row=row, column=4).number_format = DECIMAL_FORMAT
    ws.cell(row=row, column=5, value="(should = 1.5x)")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER

def create_unit_economics_sheet(ws):
    """Create Tab 2: Unit Economics"""
    ws.title = "2. Unit Economics"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    # Header
    headers = ["Metric", "Optimistic", "Base", "Pessimistic"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    row = 2
    metrics = [
        ("ARPU (VNĐ/khách/tháng)", "='1. Assumptions'!B2", "='1. Assumptions'!C2", "='1. Assumptions'!D2"),
        ("Total COGS (VNĐ/khách/tháng)", "='1. Assumptions'!B10", "='1. Assumptions'!C10", "='1. Assumptions'!D10"),
        ("Gross Profit (VNĐ/khách/tháng)", "=B2-B3", "=C2-C3", "=D2-D3"),
        ("Gross Margin %", "=B4/B2", "=C4/C2", "=D4/D2"),
        ("Monthly Churn Rate", "='1. Assumptions'!B22", "='1. Assumptions'!C22", "='1. Assumptions'!D22"),
        ("Avg Customer Lifetime (tháng)", "=1/B5", "=1/C5", "=1/D5"),
        ("LTV (triệu VNĐ/khách)", "=B4*B6", "=C4*C6", "=D4*D6"),
        ("CAC (VNĐ/khách)", "='1. Assumptions'!B25", "='1. Assumptions'!C25", "='1. Assumptions'!D25"),
        ("LTV / CAC Ratio", "=B7/B8", "=C7/C8", "=D7/D8"),
        ("CAC Payback (tháng)", "=B8/B4", "=C8/C4", "=D8/D4"),
        ("Status", "=IF(AND(B9>3, B10<12, B4>0), \"HEALTHY\", \"UNHEALTHY\")", 
         "=IF(AND(C9>3, C10<12, C4>0), \"HEALTHY\", \"UNHEALTHY\")",
         "=IF(AND(D9>3, D10<12, D4>0), \"HEALTHY\", \"UNHEALTHY\")"),
    ]
    
    for label, opt_formula, base_formula, pess_formula in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        
        for col, formula in enumerate([opt_formula, base_formula, pess_formula], start=2):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = GREEN_FONT
            cell.border = THIN_BORDER
            
            # Format
            if "Margin" in label or "Churn" in label:
                cell.number_format = PERCENT_FORMAT
            elif "Ratio" in label or "Lifetime" in label or "Payback" in label:
                cell.number_format = DECIMAL_FORMAT
            elif "Status" not in label:
                cell.number_format = CURRENCY_FORMAT if "LTV" in label or "CAC" in label else DECIMAL_FORMAT
            
            # Conditional formatting for Status
            if "Status" in label:
                cell.alignment = CENTER_ALIGN
        
        row += 1

def create_pl_roi_sheet(ws):
    """Create Tab 3: P&L & ROI with 36-month projection"""
    ws.title = "3. P&L & ROI"
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    
    # Scenario selector
    ws.cell(row=1, column=1, value="Scenario:")
    ws.cell(row=1, column=1).font = BOLD_FONT
    scenario_cell = ws.cell(row=1, column=2, value="Base")
    scenario_cell.fill = YELLOW_FILL
    scenario_cell.font = BLUE_FONT
    
    # Data validation for scenario dropdown
    dv = DataValidation(type="list", formula1='"Optimistic,Base,Pessimistic"', allow_blank=False)
    dv.error = 'Please select Optimistic, Base, or Pessimistic'
    ws.add_data_validation(dv)
    dv.add(scenario_cell)
    
    row = 3
    
    # Assumptions reference section
    ws.cell(row=row, column=1, value="Active Scenario Parameters:")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    params = [
        ("Adoption Rate (khách/tháng)", f"=IFERROR(INDEX('1. Assumptions'!$B$3:$D$3,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))/100*INDEX('1. Assumptions'!$B$4:$D$4,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0)),0)"),
        ("ARPU (M/tháng)", f"=INDEX('1. Assumptions'!$B$2:$D$2,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("COGS (M/tháng)", f"=INDEX('1. Assumptions'!$B$10:$D$10,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("Churn Rate", f"=INDEX('1. Assumptions'!$B$22:$D$22,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("CAC (M/customer)", f"=INDEX('1. Assumptions'!$B$25:$D$25,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("Fixed Cost (M/tháng)", f"=INDEX('1. Assumptions'!$B$33:$D$33,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("Initial Investment (M)", f"=INDEX('1. Assumptions'!$B$36:$D$36,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("Initial Cash (M)", f"=INDEX('1. Assumptions'!$B$37:$D$37,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
        ("Discount Rate", f"=INDEX('1. Assumptions'!$B$38:$D$38,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))"),
    ]
    
    for param_name, formula in params:
        ws.cell(row=row, column=1, value=param_name)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=2).font = GREEN_FONT
        ws.cell(row=row, column=2).number_format = DECIMAL_FORMAT
        row += 1
    
    row += 2
    
    # 36-month projection table header
    ws.cell(row=row, column=1, value="36-MONTH PROJECTION")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    headers = ["Month", "Cust Start", "New", "Churned", "Cust End", "Revenue", "COGS", "Gross Profit", 
               "S&M Cost", "Fixed Cost", "Net CF", "Cash Position"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    header_row = row
    row += 1
    
    # Month 0 (startup setup)
    ws.cell(row=row, column=1, value=0)
    # Month 0: no customers, just investment
    for col in range(2, 12):
        ws.cell(row=row, column=col, value=0)
    ws.cell(row=row, column=11, value=f"=-INDEX('1. Assumptions'!$B$36:$D$36,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))-INDEX('1. Assumptions'!$B$37:$D$37,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))")
    ws.cell(row=row, column=12, value=f"=INDEX('1. Assumptions'!$B$37:$D$37,MATCH($B$1,'1. Assumptions'!$B$1:$D$1,0))+K{row}")
    
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if col >= 5:
            ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
    
    month_zero_row = row
    row += 1
    
    # Month 1-36
    for month in range(1, 37):
        prev_row = row - 1
        
        # Month number
        ws.cell(row=row, column=1, value=month)
        
        # Customers start (previous month end)
        ws.cell(row=row, column=2, value=f"=E{prev_row}")
        
        # New customers per month (constant adoption)
        ws.cell(row=row, column=3, value=f"=$B$5" if month == 1 else f"=$B$5")  # Simplified: constant new customers
        
        # Churned customers
        ws.cell(row=row, column=4, value=f"=B{row}*$B$8")
        
        # Customers end
        ws.cell(row=row, column=5, value=f"=B{row}-D{row}+C{row}")
        
        # Revenue
        ws.cell(row=row, column=6, value=f"=E{row}*$B$6")
        
        # COGS
        ws.cell(row=row, column=7, value=f"=E{row}*$B$7")
        
        # Gross Profit
        ws.cell(row=row, column=8, value=f"=F{row}-G{row}")
        
        # S&M Cost (new customers × CAC)
        ws.cell(row=row, column=9, value=f"=C{row}*$B$9")
        
        # Fixed Cost
        ws.cell(row=row, column=10, value=f"=$B$10")
        
        # Net Cash Flow
        ws.cell(row=row, column=11, value=f"=H{row}-I{row}-J{row}")
        
        # Cash Position
        if month == 1:
            ws.cell(row=row, column=12, value=f"=L{month_zero_row}+K{row}")
        else:
            ws.cell(row=row, column=12, value=f"=L{prev_row}+K{row}")
        
        # Formatting
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if col >= 5:
                ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
                # Conditional formatting for negative cash
                if col == 12 and month % 3 == 0:  # Every 3 months for visibility
                    pass
        
        row += 1
    
    # KPI Summary Section
    row += 2
    ws.cell(row=row, column=1, value="KEY PERFORMANCE INDICATORS")
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1
    
    kpis = [
        ("Break-even Month", "=IFERROR(INDEX(A41:A76,MATCH(TRUE,(K41:K76)>0,0)),\"N/A\")"),
        ("NPV 36 months", "=IFERROR(NPV(C38/12, K41:K76),0)"),
        ("IRR Monthly %", "=IFERROR(IRR(K40:K76)*100,0)"),
        ("IRR Annualized %", "=IFERROR(((1+D20/100)^12-1)*100,0)"),
        ("Project Payback (months)", "=IFERROR(INDEX(A41:A76,MATCH(TRUE,(L41:L76)>=INDEX(L41:L76,1),0)),\"N/A\")"),
        ("Min Cash (first 12mo)", "=MIN(L41:L52)"),
        ("Cash at Month 12", "=L52"),
        ("Runway until negative", "=IFERROR(INDEX(A41:A76,MATCH(TRUE,(L41:L76)<0,0)),36)"),
    ]
    
    for kpi_name, formula in kpis:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=2).font = GREEN_FONT
        ws.cell(row=row, column=2).number_format = DECIMAL_FORMAT
        row += 1

def main():
    """Main function to generate Excel file"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create sheets
    ws1 = wb.create_sheet("1. Assumptions")
    ws2 = wb.create_sheet("2. Unit Economics")
    ws3 = wb.create_sheet("3. P&L & ROI")
    
    # Populate sheets
    create_assumptions_sheet(ws1)
    create_unit_economics_sheet(ws2)
    create_pl_roi_sheet(ws3)
    
    # Save
    output_path = Path(__file__).parent / FILE_NAME
    wb.save(output_path)
    
    print(f"✅ File created successfully: {output_path}")
    print(f"   - Tab 1: Assumptions (6 sections, 3 scenarios, yellow cells for input)")
    print(f"   - Tab 2: Unit Economics (LTV, CAC, LTV/CAC ratio, CAC Payback)")
    print(f"   - Tab 3: P&L & ROI (36-month projection, KPI summary)")

if __name__ == "__main__":
    main()
